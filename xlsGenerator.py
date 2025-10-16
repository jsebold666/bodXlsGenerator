import pandas as pd
import os
from datetime import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import re
import bs_config 

def export_to_excel(data, sheet_name="BODs", filename=None):
    if not filename:
        filename = f"bod_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    full_path = os.path.join(bs_config.export_path, filename)

    df = pd.DataFrame(data)

    # Extrai quantidade, item, material e cria coluna combinada
    df["Quantity"] = df["Bod"].str.extract(r"^(\d+)")
    df["Item"] = df["Bod"].str.extract(r"^\d+(?:verite|valorite|agapite|gold)?([a-z]+)")
    df["Item"] = df["Item"].str.upper()
    df["Material"] = df["Type"].str.upper()
    df["ColumnGroup"] = df["Material"] + " " + df["Quantity"] + "e"

    # Cria tabela pivô
    pivot_df = df.pivot_table(index="Item", columns="ColumnGroup", values="Amount", aggfunc="sum", fill_value=0)
    pivot_df = pivot_df.sort_index(axis=1)

    # NÃO separar COIF em colunas extras, mantém a linha COIF intacta

    # Reordena as linhas na ordem desejada
    row_order = ["LBOD", "COIF", "LEGS", "TUNIC", "ARMS", "GLOVES", "GORGET", "HELM"]
    pivot_df = pivot_df.reindex(row_order + [i for i in pivot_df.index if i not in row_order])

    # Estilos
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    agapite_fill = PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid")

    material_fill_map = {}
    for col in pivot_df.columns:
        base_col = col.replace("COIF ", "")
        if "VALORITE" in base_col:
            material_fill_map[col] = blue_fill
        elif "VERITE" in base_col:
            material_fill_map[col] = green_fill
        elif "AGAPITE" in base_col:
            material_fill_map[col] = agapite_fill
        elif "GOLD" in base_col:
            material_fill_map[col] = yellow_fill

    # Criar planilha
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    for r_idx, row in enumerate(dataframe_to_rows(pivot_df, index=True, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            if r_idx == 1 and c_idx > 1:
                col_name = list(pivot_df.columns)[c_idx - 2]
                if col_name in material_fill_map:
                    cell.fill = material_fill_map[col_name]

            elif isinstance(value, (int, float)) and value == 0:
                cell.fill = red_fill

            elif r_idx > 1 and c_idx > 1:
                col_name = list(pivot_df.columns)[c_idx - 2]
                if col_name in material_fill_map:
                    cell.fill = material_fill_map[col_name]

    try:
        wb.save(full_path)
        print(f"[✔] Data exported to: {full_path}")
        return True
    except Exception as e:
        print(f"[❌ Error saving file: {e}")
        return False